"""
ChronoLegal AI — South Africa Edition
Core extraction engine: PDF → DeepSeek → Discrepancy Matrix
"""

import os
import json
import time
import io
import httpx
import fitz  # PyMuPDF
from datetime import datetime

# ── DeepSeek client ────────────────────────────────────────────────────────
DEEPSEEK_BASE = "https://api.deepseek.com/v1"
DEEPSEEK_MODEL = "deepseek-chat"
MAX_RETRIES = 3
CHUNK_SIZE = 25  # pages per chunk
MAX_PAGES_TOTAL = 500

# ── SA Medico-Legal System Prompt ──────────────────────────────────────────
# Engineered by Prompt Engineer agent — Role→Constraints→Reasoning→Examples
# structure with embedded SA law context and 2 few-shot examples
SYSTEM_PROMPT = """You are ChronoLegal AI, an SA medico-legal analyst for plaintiff attorneys in RAF and medical negligence High Court matters. Analyze three documents (hospital records, Expert A report, Expert B report) and output ONLY a discrepancy matrix in valid JSON. No commentary, no markdown.

### ROLE
You act as a medico-legal expert fluent in RAF Act 56/1996 heads of damages (past/future medical, loss of earnings, general damages), Uniform Rules 36(9)(a)/(b) and 38(2), the Michael test for logical defensibility of expert evidence, and apportionment of pre-existing conditions under RAF Amendment Act 19/2005.

### CONSTRAINTS
1. Output raw JSON per the Examples schema. No fences, preamble, or trailing text.
2. Cite specific content from each source per entry. If a source is silent, write "Not addressed" — never invent data.
3. riskLevel: "Low" = trivial (typo, immaterial date). "Medium" = material, affects one head of damage. "High" = undermines causation, quantum, or witness credibility.
4. strategicAlert: One sentence citing SA law (RAF Regulation 3, Amendment Act 19/2005, Michael test, Rule 36).

### REASONING CHAIN (apply per discrepancy)
1. Identify conflict across the three sources.
2. Classify domain: GCS/clinical severity, pre-existing condition, employability, timeline.
3. Assess materiality: which head of damage impacted.
4. Determine exploitability: cross-examination weapon or plaintiff vulnerability.
5. Ground in SA law: RAF cap, apportionment, Michael test, Practice Directive.

### SA MEDICO-LEGAL KNOWLEDGE
- GCS: Hospital GCS (contemporaneous) may conflict with expert-retrospective GCS due to intubation, sedation, or alcohol. GCS 13 vs 15 changes TBI-severity classification and RAF "serious injury" threshold (Regulation 3).
- Pre-existing conditions: Distinguish vulnerability (thin-skull: defendant takes plaintiff as found) from active disability (apportionment per RAF Amendment Act 19/2005). Hospital records carry greater evidentiary weight than retrospective expert opinion on prior degeneration.
- Employability: Occupational therapists evaluate physical capacity; industrial psychologists project labour-market earnings. Job-category conflicts directly alter loss-of-earnings quantum.
- Expert partiality: The Michael test (SCA) requires opinions be "logically defensible." If an expert's conclusion contradicts hospital records without explanation, their report is vulnerable to exclusion.
- Practice: Expert summaries are exchanged per Rule 36(9)(a)/(b). Pre-trial discrepancies must be narrowed via joint minutes or risk adverse cost orders.

### EXAMPLES

Ex1 — GCS Conflict:
Input: Hospital: "GCS 15/15, no LOC, discharged." Expert A: "Retrospective GCS 13, moderate TBI, 6% WPI." Expert B: "GCS 14, mild TBI, 1% WPI."
Output: {"timeline":[{"date":"2024-06-12","eventOrInjury":"MVA — side-impact collision","hospitalRecordNote":"GCS 15/15. No LOC. Discharged same day.","expertANote":"Retrospective GCS 13/15. Moderate TBI, 6% WPI.","expertBNote":"GCS 14/15. Mild TBI, 1% WPI.","riskLevel":"High","strategicAlert":"Hospital GCS 15/15 with no LOC contradicts both TBI opinions; general damages may fail RAF Regulation 3 serious-injury threshold. Cross-examine under Michael test — contemporaneous records outweigh retrospective opinion."}]}

Ex2 — Pre-existing Condition/Employability:
Input: Hospital: "Pre-existing lumbar degeneration." Expert A (ortho): "Trauma aggravated asymptomatic condition, total disability." Expert B (OT): "Constitutional, not trauma-related; fit for sedentary work."
Output: {"timeline":[{"date":"2024-09-03","eventOrInjury":"Rear-end collision — lumbar","hospitalRecordNote":"Pre-existing lumbar degeneration on X-ray.","expertANote":"Aggravation of asymptomatic degeneration; total permanent disability.","expertBNote":"Constitutional degeneration; fit for sedentary employment.","riskLevel":"High","strategicAlert":"Apportionment risk per RAF Amendment Act 19/2005 — loss-of-earnings may reduce 50-80%. Commission joint minute under Rule 36(9)(a); consider rheumatologist tie-breaker."}]}
"""

USER_PROMPT_TEMPLATE = """Analyze the following three medico-legal documents and produce a discrepancy matrix as strict JSON.

=== HOSPITAL RECORDS ===
{hospital_record}

=== EXPERT A REPORT ===
{expert_a_report}

=== EXPERT B REPORT ===
{expert_b_report}

Produce the DiscrepancyMatrix JSON output now."""


def _call_deepseek(system_prompt: str, user_content: str) -> dict:
    """Call DeepSeek API and return parsed JSON response."""
    api_key = os.environ.get("DEEPSEEK_API_KEY")
    if not api_key:
        raise RuntimeError("DEEPSEEK_API_KEY not set in environment")

    for attempt in range(MAX_RETRIES):
        try:
            with httpx.Client(timeout=120.0) as client:
                resp = client.post(
                    f"{DEEPSEEK_BASE}/chat/completions",
                    headers={
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json",
                    },
                    json={
                        "model": DEEPSEEK_MODEL,
                        "messages": [
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": user_content},
                        ],
                        "temperature": 0.1,      # Low temp for deterministic legal analysis
                        "max_tokens": 8192,       # Large enough for detailed matrix
                        "response_format": {"type": "json_object"},
                    },
                )
                resp.raise_for_status()
                data = resp.json()
                content = data["choices"][0]["message"]["content"].strip()

                # Parse JSON — sometimes wrapped in markdown fences
                if content.startswith("```"):
                    lines = content.split("\n")
                    content = "\n".join(
                        lines[1:-1] if lines[-1].strip() == "```" else lines[1:]
                    )

                result = json.loads(content)

                # Validate structure
                if "timeline" not in result or not isinstance(result["timeline"], list):
                    raise ValueError("Response missing 'timeline' array")

                return result

        except (json.JSONDecodeError, ValueError) as e:
            if attempt < MAX_RETRIES - 1:
                time.sleep(1 * (attempt + 1))
                continue
            raise RuntimeError(f"Failed to parse DeepSeek response after {MAX_RETRIES} attempts: {e}")
        except httpx.HTTPStatusError as e:
            if attempt < MAX_RETRIES - 1:
                time.sleep(2 * (attempt + 1))
                continue
            raise RuntimeError(f"DeepSeek API error: {e.response.status_code} — {e.response.text[:300]}")
        except Exception:
            if attempt < MAX_RETRIES - 1:
                time.sleep(2 * (attempt + 1))
                continue
            raise

    raise RuntimeError("DeepSeek call failed after all retries")


def extract_text_from_pdf(pdf_bytes: bytes, filename: str, max_pages: int = 200) -> list[tuple[str, str, int]]:
    """Extract text from a PDF, returning list of (filename, page_text, page_number)."""
    pages = []
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    total = min(len(doc), max_pages)

    for page_num in range(total):
        text = doc[page_num].get_text()
        if text.strip():
            pages.append((filename, text.strip(), page_num + 1))
    doc.close()
    return pages


def build_document_text(pages: list[tuple[str, str, int]], label: str) -> str:
    """Build paginated text block for a document."""
    if not pages:
        return "[No text could be extracted from this document]"

    lines = []
    total_chars = 0
    max_chars = 25000  # Cap to avoid exceeding context window

    for filename, text, page_num in pages:
        header = f"\n--- {label} | Page {page_num} ---\n"
        if total_chars + len(header) + len(text) > max_chars:
            remaining = max_chars - total_chars - len(header)
            if remaining > 200:
                lines.append(header + text[:remaining] + "\n[... truncated — page limit reached]")
            break
        lines.append(header + text)
        total_chars += len(header) + len(text)

    return "\n".join(lines)


def analyze_case_bundle(
    hospital_pdf: tuple[bytes, str],
    expert_a_pdf: tuple[bytes, str],
    expert_b_pdf: tuple[bytes, str],
    max_pages_per_doc: int = 200,
) -> dict:
    """
    Main entry point: takes 3 PDFs, returns DiscrepancyMatrix.
    
    Returns dict with:
    - timeline: the discrepancy matrix
    - metadata: processing info
    """
    # Extract text from each PDF
    hospital_pages = extract_text_from_pdf(hospital_pdf[0], hospital_pdf[1], max_pages_per_doc)
    expert_a_pages = extract_text_from_pdf(expert_a_pdf[0], expert_a_pdf[1], max_pages_per_doc)
    expert_b_pages = extract_text_from_pdf(expert_b_pdf[0], expert_b_pdf[1], max_pages_per_doc)

    total_pages = len(hospital_pages) + len(expert_a_pages) + len(expert_b_pages)
    if total_pages == 0:
        raise ValueError("No extractable text found in any of the uploaded PDFs")

    # Build document blocks
    hospital_text = build_document_text(hospital_pages, "HOSPITAL RECORD")
    expert_a_text = build_document_text(expert_a_pages, "EXPERT REPORT A")
    expert_b_text = build_document_text(expert_b_pages, "EXPERT REPORT B")

    # Format user prompt
    user_prompt = USER_PROMPT_TEMPLATE.format(
        hospital_record=hospital_text,
        expert_a_report=expert_a_text,
        expert_b_report=expert_b_text,
    )

    # Call DeepSeek
    start_time = time.time()
    result = _call_deepseek(SYSTEM_PROMPT, user_prompt)
    elapsed = time.time() - start_time

    # Add metadata
    result["metadata"] = {
        "model": DEEPSEEK_MODEL,
        "processedAt": datetime.utcnow().isoformat() + "Z",
        "processingTimeSeconds": round(elapsed, 2),
        "totalPagesAnalyzed": total_pages,
        "hospitalPages": len(hospital_pages),
        "expertAPages": len(expert_a_pages),
        "expertBPages": len(expert_b_pages),
        "hospitalFilename": hospital_pdf[1],
        "expertAFilename": expert_a_pdf[1],
        "expertBFilename": expert_b_pdf[1],
    }

    return result


# ── Excel Export (preserved from original, adapted for matrix) ─────────────

def generate_matrix_excel(result: dict, case_name: str) -> io.BytesIO:
    """Generate Excel workbook from discrepancy matrix."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    timeline = result.get("timeline", [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Discrepancy Matrix"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    risk_colors = {"High": "C00000", "Medium": "E36C09", "Low": "4CAF50"}
    risk_font_colors = {"High": "FFFFFF", "Medium": "FFFFFF", "Low": "FFFFFF"}

    headers = ["Date", "Event / Injury", "Hospital Record", "Expert A Note", "Expert B Note", "Risk", "Strategic Alert"]
    col_widths = [12, 22, 30, 30, 30, 10, 45]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    for idx, event in enumerate(timeline, 2):
        risk = event.get("riskLevel", "Low")
        row_data = [
            event.get("date", ""),
            event.get("eventOrInjury", ""),
            event.get("hospitalRecordNote", ""),
            event.get("expertANote", ""),
            event.get("expertBNote", ""),
            risk,
            event.get("strategicAlert", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 6:  # Risk level
                cell.fill = PatternFill(
                    start_color=risk_colors.get(risk, "A6A6A6"),
                    end_color=risk_colors.get(risk, "A6A6A6"),
                    fill_type="solid",
                )
                cell.font = Font(bold=True, color=risk_font_colors.get(risk, "FFFFFF"))
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Summary sheet
    sw = wb.create_sheet("Summary")
    sw["A1"] = "ChronoLegal AI — Discrepancy Matrix Report"
    sw["A1"].font = Font(bold=True, size=14, color="1F4E79")
    meta = [
        ("Case", case_name),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Timeline Entries", len(timeline)),
        ("High Risk Items", sum(1 for e in timeline if e.get("riskLevel") == "High")),
        ("Medium Risk Items", sum(1 for e in timeline if e.get("riskLevel") == "Medium")),
        ("Low Risk Items", sum(1 for e in timeline if e.get("riskLevel") == "Low")),
    ]
    for row_num, (label, value) in enumerate(meta, 3):
        sw.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        sw.cell(row=row_num, column=2, value=str(value))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
